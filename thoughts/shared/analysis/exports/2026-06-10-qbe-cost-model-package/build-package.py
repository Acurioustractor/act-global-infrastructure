#!/usr/bin/env python3
"""Build the QBE cost-model send package (xlsx + csv) from the canonical
explorer JSON (Goods Asset Register/v2 cost-model-scenarios.json, meta.version v6).
Re-run after any v7 update: python3 build-package.py
"""
import csv
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = "/Users/benknight/Code/Goods Asset Register/v2/src/lib/data/cost-model-scenarios.json"
OUT = os.path.dirname(os.path.abspath(__file__))
CSV_DIR = os.path.join(OUT, "csv")
os.makedirs(CSV_DIR, exist_ok=True)

d = json.load(open(SRC))
assert d["meta"]["version"] == "v6", f"expected v6, got {d['meta']['version']} — review README before re-running"

# ---------- styles ----------
H = Font(bold=True, size=12)
HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="121212")
INPUT_FILL = PatternFill("solid", fgColor="FFF3B0")   # editable assumption
VERIFIED_FILL = PatternFill("solid", fgColor="D9EAD3")
UNVERIFIED_FILL = PatternFill("solid", fgColor="F4CCCC")
NOTE = Font(italic=True, size=9, color="555555")
MONEY = "#,##0.00"
MONEY0 = "#,##0"

def sheet(wb, title):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = True
    return ws

def header_row(ws, row, cols, widths=None):
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HDR
        cell.fill = HDR_FILL
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def write_csv(name, rows):
    with open(os.path.join(CSV_DIR, name), "w", newline="") as f:
        csv.writer(f).writerows(rows)

wb = Workbook()
wb.remove(wb.active)

# ---------- READ ME ----------
ws = sheet(wb, "READ ME")
ws.column_dimensions["A"].width = 110
lines = [
    ("Goods on Country — bed cost model v6 (worked numbers)", H),
    (f"Source of truth: cost-model explorer JSON, version {d['meta']['version']} (supersedes {d['meta'].get('supersedes')}), created {d['meta'].get('created')}", None),
    ("Prepared for: Matt Allen (Social Impact Hub) / Malcolm Aikman (QA) — QBE Catalysing Impact 2026, Stage 2 prep", None),
    ("", None),
    ("CONFIDENCE TAGS used throughout:", H),
    ("  VERIFIED — invoice / first-party document (Xero, Defy invoices, supplier quotes)", None),
    ("  INFERRED — computed from verified inputs + a stated assumption", None),
    ("  UNVERIFIED — needs a vendor quote or a founder decision; do not anchor on it", None),
    ("", None),
    ("KNOWN LIMITS (stated up front, in QBE claim discipline):", H),
    ("  1. 3-site standalone capex ($300–450K) is UNVERIFIED — vendor quotes pending (the gating item).", None),
    ("  2. Debt-service table is INTEREST-ONLY. Sheet 6 adds an amortising view: at $500K/5yr the annual obligation", None),
    ("     (~$126K) exceeds base brokerage income (~$96K) — repayable matching capital needs patient/bullet terms,", None),
    ("     a longer tenor, or a smaller debt slice. We flag this rather than have you find it.", None),
    ("  3. Freight base $100/bed is an estimate; verified Defy pallet lines imply $101–185/bed Botany→Alice (8/pallet).", None),
    ("  4. Per-site throughput 250 beds/yr is a modelling assumption, not a demonstrated rate.", None),
    ("  5. Corporate-RAP gift price ($800) is the biggest revenue unknown — no firm anchor yet.", None),
    ("  6. Management data, not audited. Xero 'actual cost' extracts must follow the mis-tag cleanup (sheet 8).", None),
    ("  7. Year-1 utilisation 20% assumes 3 simultaneous site openings; sequential opening tells the same totals", None),
    ("     with a more defensible per-site story.", None),
    ("", None),
    ("EDITABLE ASSUMPTION CELLS are yellow. Change them and the dependent figures recalculate.", None),
]
for i, (text, font) in enumerate(lines, start=1):
    c = ws.cell(row=i, column=1, value=text)
    if font:
        c.font = font

# ---------- 1 Build states ----------
ws = sheet(wb, "1 Build states")
header_row(ws, 1, ["State", "Component", "$ / bed", "Direct total", "Capex added", "Capex cumulative", "Beds/day"], [34, 44, 12, 14, 14, 18, 10])
r = 2
csv_rows = [["State", "Component", "$/bed", "Direct total", "Capex added", "Capex cumulative", "Beds/day"]]
for key, st in d["build_states"].items():
    first = r
    for comp in st["components"]:
        ws.cell(row=r, column=1, value=st["label"] if r == first else None)
        ws.cell(row=r, column=2, value=comp["label"])
        cell = ws.cell(row=r, column=3, value=comp["amount"])
        cell.number_format = MONEY
        cell.fill = INPUT_FILL
        csv_rows.append([st["label"], comp["label"], comp["amount"], "", "", "", ""])
        r += 1
    tot = ws.cell(row=first, column=4, value=f"=SUM(C{first}:C{r-1})")
    tot.number_format = MONEY
    tot.font = Font(bold=True)
    ws.cell(row=first, column=5, value=st.get("capital_added")).number_format = MONEY0
    ws.cell(row=first, column=6, value=st.get("capital_cumulative")).number_format = MONEY0
    ws.cell(row=first, column=7, value=st.get("throughput_beds_per_day"))
    csv_rows.append([st["label"], "DIRECT TOTAL", st["direct_total"], st["direct_total"], st.get("capital_added"), st.get("capital_cumulative"), st.get("throughput_beds_per_day")])
    r += 1
ws.cell(row=r + 1, column=1, value="Component costs are editable (yellow); totals are live SUM formulas.").font = NOTE
write_csv("01-build-states.csv", csv_rows)

# ---------- 2 Verified rates ----------
ws = sheet(wb, "2 Verified rates")
header_row(ws, 1, ["Item", "Rate ($)", "Source", "Confidence"], [42, 12, 50, 14])
csv_rows = [["Item", "Rate", "Source", "Confidence"]]
r = 2
for key, v in d["defy_verified_rates"].items():
    if not isinstance(v, dict):
        continue
    ws.cell(row=r, column=1, value=key)
    ws.cell(row=r, column=2, value=v.get("amount", v.get("rate"))).number_format = MONEY
    ws.cell(row=r, column=3, value=v.get("source", ""))
    conf = v.get("confidence", "")
    c = ws.cell(row=r, column=4, value=conf)
    if conf == "verified":
        c.fill = VERIFIED_FILL
    csv_rows.append([key, v.get("amount", v.get("rate")), v.get("source", ""), conf])
    r += 1
for key, val in d["physics"].items():
    if key.startswith("_"):
        continue
    ws.cell(row=r, column=1, value=f"physics: {key}")
    ws.cell(row=r, column=2, value=val)
    csv_rows.append([f"physics:{key}", val, "", ""])
    r += 1
for key, val in d["labour_rates_in_house"].items():
    if key.startswith("_"):
        continue
    ws.cell(row=r, column=1, value=f"labour: {key}")
    ws.cell(row=r, column=2, value=val)
    csv_rows.append([f"labour:{key}", val, "", ""])
    r += 1
write_csv("02-verified-rates.csv", csv_rows)

# ---------- 3 Volume & fully-loaded ----------
ws = sheet(wb, "3 Volume & fully-loaded")
header_row(ws, 1, ["Volume", "S2 Defy kits", "S3 Defy panels", "S4 Factory", "S5 Community"], [22, 14, 14, 14, 14])
csv_rows = [["Volume", "S2", "S3", "S4", "S5"]]
r = 2
for row in d["fully_loaded_grid"]:
    ws.cell(row=r, column=1, value=row["volume_label"])
    for i, k in enumerate(["state_2_defy_kits", "state_3_defy_panels", "state_4_factory", "state_5_community"], start=2):
        ws.cell(row=r, column=i, value=row[k]).number_format = MONEY0
    csv_rows.append([row["volume_label"], row["state_2_defy_kits"], row["state_3_defy_panels"], row["state_4_factory"], row["state_5_community"]])
    r += 1
ws.cell(row=r, column=1, value="Fully-loaded = direct + founder production overhead + admin + field travel + freight (volume-scaled).").font = NOTE
r += 2
ws.cell(row=r, column=1, value="v6 3-site volume ramp (Decision 8)").font = H
r += 1
header_row(ws, r, ["Year", "Beds", "Per site (÷3)", "Utilisation %"])
ramp = d["volume_ramp_v6"]
csv_rows.append([])
csv_rows.append(["Year", "Beds", "Per site", "Utilisation %"])
for i, y in enumerate(["year_1", "year_2", "year_3"], start=1):
    r += 1
    yr = ramp[y]
    ws.cell(row=r, column=1, value=f"Year {i}")
    ws.cell(row=r, column=2, value=yr["beds"])
    ws.cell(row=r, column=3, value=yr["per_site"])
    ws.cell(row=r, column=4, value=yr["utilisation_pct"])
    csv_rows.append([f"Year {i}", yr["beds"], yr["per_site"], yr["utilisation_pct"]])
r += 1
ws.cell(row=r, column=1, value=f"Per-site capacity {ramp['per_site_capacity_per_year']}/yr is a MODELLING ASSUMPTION.").font = NOTE
ws.cell(row=r, column=1).fill = UNVERIFIED_FILL
write_csv("03-volume-fully-loaded.csv", csv_rows)

# ---------- 4 Site capex ----------
ws = sheet(wb, "4 Site capex (UNVERIFIED)")
cap = d["standalone_site_capex"]
header_row(ws, 1, ["", "Low", "Medium", "High"], [30, 14, 14, 14])
rows = [
    ("Per-site kit", cap["per_site_low"], cap["per_site_med"], cap["per_site_high"]),
    (f"× {cap['sites_modelled']} sites", cap["three_sites_low"], cap["three_sites_med"], cap["three_sites_high"]),
]
csv_rows = [["", "Low", "Medium", "High"]]
for i, (label, lo, med, hi) in enumerate(rows, start=2):
    ws.cell(row=i, column=1, value=label)
    for j, v in enumerate([lo, med, hi], start=2):
        c = ws.cell(row=i, column=j, value=v)
        c.number_format = MONEY0
        c.fill = UNVERIFIED_FILL
    csv_rows.append([label, lo, med, hi])
ws.cell(row=5, column=1, value=cap["_note"]).font = NOTE
csv_rows.append([cap["_note"]])
write_csv("04-site-capex.csv", csv_rows)

# ---------- 5 Demand & revenue (live formulas) ----------
ws = sheet(wb, "5 Demand & revenue")
header_row(ws, 1, ["Segment", "Share %", "Price/bed (edit)", "Confidence", "Y1 beds", "Y2 beds", "Y3 beds", "3yr revenue"], [28, 10, 16, 14, 10, 10, 10, 14])
years = [d["volume_ramp_v6"][y]["beds"] for y in ["year_1", "year_2", "year_3"]]
csv_rows = [["Segment", "Share %", "Price/bed", "Confidence", "Y1 beds", "Y2 beds", "Y3 beds", "3yr revenue"]]
r = 2
for seg in d["demand_mix"]["segments"]:
    ws.cell(row=r, column=1, value=seg["label"])
    ws.cell(row=r, column=2, value=seg["share_pct"])
    p = ws.cell(row=r, column=3, value=seg["price_per_bed"])
    p.number_format = MONEY0
    p.fill = INPUT_FILL
    conf = seg.get("confidence", "")
    c = ws.cell(row=r, column=4, value=conf)
    c.fill = VERIFIED_FILL if conf == "verified" else (UNVERIFIED_FILL if conf == "unverified" else PatternFill())
    beds = [round(y * seg["share_pct"] / 100) for y in years]
    for j, b in enumerate(beds, start=5):
        ws.cell(row=r, column=j, value=b)
    rev = ws.cell(row=r, column=8, value=f"=C{r}*SUM(E{r}:G{r})")
    rev.number_format = MONEY0
    csv_rows.append([seg["label"], seg["share_pct"], seg["price_per_bed"], conf, *beds, seg["price_per_bed"] * sum(beds)])
    r += 1
tot = ws.cell(row=r, column=8, value=f"=SUM(H2:H{r-1})")
tot.number_format = MONEY0
tot.font = Font(bold=True)
ws.cell(row=r, column=1, value="3-YEAR TOTAL").font = Font(bold=True)
ws.cell(row=r + 1, column=1, value=f"Band: ${d['demand_mix']['three_year_revenue_band_low']:,} – ${d['demand_mix']['three_year_revenue_band_high']:,} (gift price is the biggest unknown)").font = NOTE
csv_rows.append(["3-YEAR TOTAL", "", "", "", "", "", "", d["demand_mix"]["three_year_revenue"]])
write_csv("05-demand-revenue.csv", csv_rows)

# ---------- 6 Waterfall & debt (live formulas) ----------
ws = sheet(wb, "6 Waterfall & debt")
wf, ds = d["margin_waterfall"], d["debt_service"]
ws.column_dimensions["A"].width = 40
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 70
inputs = [
    ("Sale price ($/bed)", wf["sale_price"], "verified anchor: Centrecorp INV-0291 $801"),
    ("COGS — community state ($/bed)", wf["cogs_community"], "inferred (sheet 1, State 5)"),
    ("Freight ($/bed)", wf["freight"], "ESTIMATE — verified pallet lines imply $101–185 Botany→Alice"),
    ("ACT brokerage ($/bed)", wf["act_brokerage_default"], f"slider {wf['act_brokerage_min']}–{wf['act_brokerage_max']}; must flex DOWN in a downside (floor ~${d['sensitivity_v6']['joint_downside']['brokerage_flex_floor_per_bed']})"),
    ("Cashflow buffer ($/bed)", wf["cashflow_buffer"], "4-state buffer policy"),
    ("Revenue beds / yr (Y3, 80% of vol)", ds["revenue_beds_per_year"], "institutional + gift segments only"),
    ("Debt drawn ($)", ds["debt_high"], f"range {ds['debt_low']:,}–{ds['debt_high']:,}"),
    ("Interest rate (%)", ds["interest_rate_high_pct"], f"range {ds['interest_rate_low_pct']}–{ds['interest_rate_high_pct']}%"),
    ("Amortisation term (yrs, 0 = interest-only)", 0, "v6 tables are interest-only; set 3 or 5 to see amortising reality"),
]
ws.cell(row=1, column=1, value="INPUTS (yellow — edit these)").font = H
r = 2
for label, val, note in inputs:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=val)
    c.fill = INPUT_FILL
    c.number_format = MONEY0 if val and val > 100 else "0"
    ws.cell(row=r, column=3, value=note).font = NOTE
    r += 1
ws.cell(row=r + 1, column=1, value="PER-BED WATERFALL (computed)").font = H
calc_start = r + 2
calcs = [
    ("Landed margin ($/bed)", "=B2-B3-B4"),
    ("Community co-op margin ($/bed)", "=B2-B3-B4-B5-B6"),
    ("Community total benefit ($/bed, incl labour)", "=130+B2-B3-B4-B5-B6"),
    ("ACT brokerage income ($/yr)", "=B5*B7"),
    ("Interest-only debt service ($/yr)", "=B8*B9/100"),
    ("Amortising payment ($/yr, term>0)", "=IF(B10=0, B8*B9/100, B8*(B9/100)/(1-(1+B9/100)^-B10))"),
    ("Coverage — interest-only (×)", "=(B5*B7)/(B8*B9/100)"),
    ("Coverage — amortising (×)", "=IF(B10=0,\"n/a\",(B5*B7)/(B8*(B9/100)/(1-(1+B9/100)^-B10)))"),
]
r = calc_start
for label, formula in calcs:
    ws.cell(row=r, column=1, value=label)
    c = ws.cell(row=r, column=2, value=formula)
    c.number_format = MONEY
    c.font = Font(bold=True)
    r += 1
ws.cell(row=r + 1, column=1, value="FINDING: at $500K amortised over 5 yrs the payment (~$122K/yr) exceeds base brokerage income ($96K/yr).").font = NOTE
ws.cell(row=r + 2, column=1, value="Repayable matching capital therefore needs patient/bullet terms, a longer tenor, or a smaller debt slice.").font = NOTE
write_csv("06-waterfall-debt.csv", [
    ["Input", "Base", "Note"], *[[a, b, c] for a, b, c in inputs],
    [],
    ["Computed (base case)", "", ""],
    ["Landed margin/bed", wf["landed_margin"], ""],
    ["Community co-op margin/bed", wf["community_coop_margin"], ""],
    ["Community total benefit/bed", wf["community_total_benefit_per_bed"], ""],
    ["ACT brokerage income/yr", ds["base_brokerage_annual"], ""],
    ["Coverage interest-only", f"{ds['coverage_low_x']}–{ds['coverage_high_x']}x", ""],
    ["Coverage $500K 5yr amortising", "~0.8x — UNDER 1. Needs patient/bullet terms.", ""],
])

# ---------- 7 Sensitivity ----------
ws = sheet(wb, "7 Sensitivity")
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 90
jd = d["sensitivity_v6"]["joint_downside"]
rows = [
    ("Dials (±20%)", ", ".join(d["sensitivity_v6"]["dials_plus_minus_20pct"])),
    ("Defy kit dial → community state", d["sensitivity_v6"]["defy_kit_dial_effect_on_community"] + " — free collected plastic decouples community COGS from Defy pricing"),
    ("", ""),
    ("3-WAY JOINT DOWNSIDE (load-bearing)", f"volume {jd['drivers']['volume_pct']}%, institutional price {jd['drivers']['institutional_price_pct']}%, labour +{jd['drivers']['community_labour_pct']}%"),
    ("Community co-op margin under downside", f"~${jd['community_coop_margin_per_bed']}/bed (collapses)"),
    ("ACT debt still serviced", str(jd["act_debt_still_serviced"])),
    ("Implication", jd["implication"]),
]
r = 1
for a, b in rows:
    ws.cell(row=r, column=1, value=a).font = Font(bold=bool(a and a == a.upper() or "DOWNSIDE" in a))
    cell = ws.cell(row=r, column=2, value=b)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
write_csv("07-sensitivity.csv", [["Item", "Value"], *rows])

# ---------- 8 Open questions & data hygiene ----------
ws = sheet(wb, "8 Open items")
ws.column_dimensions["A"].width = 56
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 24
ws.cell(row=1, column=1, value="Open questions for Defy (supplier)").font = H
r = 2
csv_rows = [["Open questions for Defy"]]
for q in d.get("open_questions_for_defy", []):
    text = q if isinstance(q, str) else json.dumps(q)
    ws.cell(row=r, column=1, value=text).alignment = Alignment(wrap_text=True)
    csv_rows.append([text])
    r += 1
r += 1
ws.cell(row=r, column=1, value="Xero mis-tags to fix BEFORE any 'actual cost' extract").font = H
csv_rows.append([])
csv_rows.append(["Supplier / line", "Amount", "Current tag", "Correct tag", "Note"])
r += 1
for m in d.get("mistags_to_fix", []):
    label = f"{m.get('supplier')} ({m.get('current_tag')} → {m.get('correct_tag')})"
    ws.cell(row=r, column=1, value=label).alignment = Alignment(wrap_text=True)
    ws.cell(row=r, column=2, value=m.get("amount")).number_format = MONEY0
    ws.cell(row=r, column=3, value=m.get("correct_tag"))
    csv_rows.append([m.get("supplier"), m.get("amount"), m.get("current_tag"), m.get("correct_tag"), m.get("_note", "")])
    r += 1
write_csv("08-open-items.csv", csv_rows)

wb.save(os.path.join(OUT, "goods-cost-model-v6.xlsx"))
print("WROTE", os.path.join(OUT, "goods-cost-model-v6.xlsx"))
print("WROTE", CSV_DIR, "(8 csvs)")
